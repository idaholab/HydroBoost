# Copyright 2025, Battelle Energy Alliance, LLC, ALL RIGHTS RESERVED

"""
Generate O&M Cost Calculator

This script imports Operation & Maintenance cost data from Excel input files
and calculates actual O&M costs based on simulation results.

Usage:
    python Generate_OM_cost.py <filename>
    
Example:
    python Generate_OM_cost.py Model_A
    (automatically resolves to Input_Spreadsheets/Model_A.xlsx)
"""

import os
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import Set, Tuple, List, Dict
from Core_Models.Price_Forecasting.helper_functions import parse_CLI_arguments


def construct_filepath(filename: str) -> str:
    """
    Construct full filepath from base filename.
    
    Args:
        filename (str): Base filename (e.g., 'Model_A')
        
    Returns:
        str: Full filepath (e.g., 'Input_Spreadsheets/Model_A.xlsx')
    """
    return os.path.join("Input_Spreadsheets", f"{filename}.xlsx")


def get_reservoir_info(filepath: str) -> List[Tuple[str, float, float]]:
    """
    Extract reservoir information from Hydro Reservoir sheet.
    Returns Reservoir_ID, Min_Volume, and Max_Volume for all reservoirs.
    
    Args:
        filepath (str): Path to Excel file
        
    Returns:
        List[Tuple[str, float, float]]: List of (reservoir_id, min_volume, max_volume)
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        if "Hydro Reservoir" not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb["Hydro Reservoir"]
        reservoirs = []
        
        # Read columns A, B, C starting from row 3 (skip headers in rows 1-2)
        for row in ws.iter_rows(min_row=3, max_col=3, values_only=True):
            reservoir_id = row[0]
            min_volume = row[1]
            max_volume = row[2]
            
            # Only add if reservoir_id is not None and not empty
            if reservoir_id and str(reservoir_id).strip():
                # Handle None values for volumes
                min_vol = float(min_volume) if min_volume is not None else 0.0
                max_vol = float(max_volume) if max_volume is not None else 0.0
                reservoirs.append((str(reservoir_id).strip(), min_vol, max_vol))
        
        wb.close()
        return reservoirs
        
    except Exception as e:
        print(f"Error reading Hydro Reservoir info: {str(e)}")
        return []


def detect_model_type(filepath: str) -> str:
    """
    Detect the model type based on file contents.
    
    Logic:
    1. Check O&M Cost sheet for "Pumped Storage Unit(s)" → Model C
    2. Else check Hydro Reservoir sheet cell C3 > 0 → Model B
    3. Else → Model A
    
    Args:
        filepath (str): Path to Excel file
        
    Returns:
        str: Model type ("Run-of-River", "Reservoir", or "Pumped Storage")
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        # Step 1: Check for Pumped Storage Unit(s) in O&M Cost sheet
        if "O&M Cost" in wb.sheetnames:
            ws = wb["O&M Cost"]
            
            # Search for "Pumped Storage Unit(s)" in the sheet
            for row in ws.iter_rows(values_only=True):
                if row[0] and "Pumped Storage Unit" in str(row[0]):
                    wb.close()
                    return "Pumped Storage"
        
        # Step 2: Check Hydro Reservoir C3
        if "Hydro Reservoir" in wb.sheetnames:
            ws = wb["Hydro Reservoir"]
            c3_value = ws["C3"].value
            
            wb.close()
            
            if c3_value is not None and float(c3_value) > 0:
                return "Reservoir"
            else:
                return "Run-of-River"
        
        # Default if sheet not found
        wb.close()
        return "Run-of-River"
        
    except Exception as e:
        print(f"Error detecting model type: {str(e)}")
        sys.exit(1)


def get_gen_tech_hydro_ids(filepath: str) -> Set[str]:
    """
    Extract all Tech_IDs from Gen Technology - Hydro sheet.
    
    Args:
        filepath (str): Path to Excel file
        
    Returns:
        Set[str]: Set of Tech_IDs (both hydro_t* and hydro_p*)
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        if "Gen Technology - Hydro" not in wb.sheetnames:
            wb.close()
            return set()
        
        ws = wb["Gen Technology - Hydro"]
        tech_ids = set()
        
        # Find Tech_ID column (should be column A)
        # Start from row 2 to skip header
        for row in ws.iter_rows(min_row=2, values_only=True):
            tech_id = row[0]
            if tech_id and isinstance(tech_id, str) and tech_id.strip():
                tech_id = tech_id.strip()
                # Skip header values and only include valid hydro Tech_IDs
                if tech_id.startswith("hydro_") and tech_id != "Tech_ID":
                    tech_ids.add(tech_id)
        
        wb.close()
        return tech_ids
        
    except Exception as e:
        print(f"Error reading Gen Technology - Hydro: {str(e)}")
        sys.exit(1)


def get_om_cost_tech_ids(filepath: str) -> Set[str]:
    """
    Extract all Tech_IDs from O&M Cost sheet.
    Handles variable number of rows in each section.
    
    Sections:
    - Hydro Unit(s)
    - Pumped Storage Unit(s) (if present)
    - Battery Unit(s)
    
    Args:
        filepath (str): Path to Excel file
        
    Returns:
        Set[str]: Set of all Tech_IDs defined in O&M Cost sheet
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        if "O&M Cost" not in wb.sheetnames:
            wb.close()
            return set()
        
        ws = wb["O&M Cost"]
        tech_ids = set()
        
        # Iterate through all rows to find Tech_IDs
        in_data_section = False
        
        for row in ws.iter_rows(values_only=True):
            cell_value = row[0] if row[0] is not None else ""
            cell_str = str(cell_value).strip()
            
            # Check if we're entering a data section
            # Headers are "Hydro Unit(s)", "Pumped Storage Unit(s)", "Battery Unit(s)"
            if "Unit(s)" in cell_str or cell_str == "Tech_ID":
                in_data_section = True
                continue
            
            # Find a valid Tech_ID
            if in_data_section and cell_str and not cell_str.startswith("$"):
                # Valid Tech_IDs: not "Tech_ID" header, not containing "Unit(s)", not containing "cost"
                if (cell_str != "Tech_ID" and 
                    "Unit(s)" not in cell_str and
                    "cost" not in cell_str.lower()):
                    tech_ids.add(cell_str)
        
        wb.close()
        return tech_ids
        
    except Exception as e:
        print(f"Error reading O&M Cost sheet: {str(e)}")
        sys.exit(1)


def get_rough_zone_settings(filepath: str) -> Tuple[bool, int, int]:
    """
    Extract rough zone settings from Simulation Setting sheet.
    
    Args:
        filepath (str): Path to Excel file
        
    Returns:
        Tuple[bool, int, int]: (rough_zone_flag, rough_zone_segment_number, num_segments)
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        if "Simulation Setting" not in wb.sheetnames:
            wb.close()
            return False, 0, 0
        
        ws = wb["Simulation Setting"]
        
        rough_zone_flag = False
        rough_zone_segment_number = 0
        num_segments = 2  # Default value
        
        # Read settings from the sheet (Setting column A, Value column B)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                setting_name = str(row[0]).strip()
                setting_value = row[1]
                
                if setting_name == "hydropower_rough_zone_flag":
                    if setting_value is not None:
                        rough_zone_flag = str(setting_value).lower() in ['true', '1', 'yes']
                
                elif setting_name == "hydropower_rough_zone_segment_number_value":
                    if setting_value is not None:
                        try:
                            rough_zone_segment_number = int(setting_value)
                        except (ValueError, TypeError):
                            rough_zone_segment_number = 0
                
                elif setting_name == "num_hydropower_performance_segment_value":
                    if setting_value is not None:
                        try:
                            num_segments = int(setting_value)
                        except (ValueError, TypeError):
                            num_segments = 2
        
        wb.close()
        return rough_zone_flag, rough_zone_segment_number, num_segments
        
    except Exception as e:
        print(f"Error reading rough zone settings: {str(e)}")
        return False, 0, 2


def get_om_cost_rates_per_tech_id(filepath: str) -> Dict[str, Dict[str, float]]:
    """
    Extract O&M cost rates for each Tech_ID from O&M Cost sheet.
    
    Returns a dictionary mapping Tech_ID to its cost parameters:
    - For hydro/pumped storage: startup_cost, shutdown_cost, ramp_up_cost, ramp_down_cost, rough_zone_cost
    - For battery: cost_per_cycle
    
    Args:
        filepath (str): Path to Excel file
        
    Returns:
        Dict[str, Dict[str, float]]: Nested dictionary with Tech_ID -> {cost_type: value}
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        
        if "O&M Cost" not in wb.sheetnames:
            wb.close()
            return {}
        
        ws = wb["O&M Cost"]
        cost_rates = {}
        
        current_section = None
        cost_headers = []
        
        for row in ws.iter_rows(values_only=True):
            cell_value = row[0] if row[0] is not None else ""
            cell_str = str(cell_value).strip()
            
            # Detect section headers
            if "Hydro Unit(s)" in cell_str:
                current_section = "hydro"
                continue
            elif "Pumped Storage Unit(s)" in cell_str:
                current_section = "pumped"
                continue
            elif "Battery Unit(s)" in cell_str:
                current_section = "battery"
                continue
            
            # Detect cost parameter headers (Tech_ID row)
            if cell_str == "Tech_ID":
                cost_headers = [str(cell).strip() if cell else "" for cell in row[1:]]
                continue
            
            # Extract cost data for each Tech_ID
            if current_section and cell_str and not cell_str.startswith("$"):
                # Valid Tech_IDs: not "Tech_ID" header, not containing "Unit(s)", not containing "cost"
                if (cell_str != "Tech_ID" and 
                    "Unit(s)" not in cell_str and
                    "cost" not in cell_str.lower()):
                    tech_id = cell_str
                    cost_data = {}
                
                    # Extract cost values for this Tech_ID
                    for i, header in enumerate(cost_headers):
                        if header and i + 1 < len(row):
                            value = row[i + 1]
                            if value is not None:
                                try:
                                    cost_data[header] = float(value)
                                except (ValueError, TypeError):
                                    cost_data[header] = 0.0
                    
                    cost_rates[tech_id] = cost_data
        
        wb.close()
        return cost_rates
        
    except Exception as e:
        print(f"Error extracting O&M cost rates: {str(e)}")
        return {}


def validate_tech_ids(gen_tech_ids: Set[str], om_cost_ids: Set[str], tech_type: str) -> Tuple[bool, List[str]]:
    """
    Validate that all Gen Technology Tech_IDs are defined in O&M Cost sheet.
    
    Args:
        gen_tech_ids (Set[str]): Tech_IDs from Gen Technology sheet
        om_cost_ids (Set[str]): Tech_IDs from O&M Cost sheet
        tech_type (str): Type description for error messages (e.g., "Hydro", "BESS")
        
    Returns:
        Tuple[bool, List[str]]: (validation_passed, list_of_missing_ids)
    """
    missing_ids = sorted(list(gen_tech_ids - om_cost_ids))
    return len(missing_ids) == 0, missing_ids


def construct_dispatch_filepath(filename: str) -> str:
    """
    Construct dispatch CSV filepath from base filename.
    
    Args:
        filename (str): Base filename (e.g., 'Model_A')
        
    Returns:
        str: Full dispatch CSV filepath
    """
    return os.path.join("Simulation_Results", filename, f"ALEAF_HydroBoost_{filename}__hydro_dispatch.csv")


def import_hydro_dispatch_data(filepath: str) -> pd.DataFrame:
    """
    Import hydro dispatch data from CSV file.
    Uses chunking for efficient reading of large files.
    
    Args:
        filepath (str): Path to the hydro dispatch CSV file
        
    Returns:
        pd.DataFrame: Hydro dispatch data
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Hydro dispatch file not found: {filepath}")
    
    try:
        dispatch_data = pd.read_csv(filepath)
        print(f"  Imported {len(dispatch_data)} rows from dispatch CSV")
        return dispatch_data
    except Exception as e:
        raise Exception(f"Error reading hydro dispatch file: {str(e)}")


def calculate_hydro_generator_startup_shutdown_events(dispatch_data: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate startup and shutdown events based on commitment transitions.
    Only processes hydro generator units (hydro_t*), ignoring pump units (hydro_p*).
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data
        
    Returns:
        pd.DataFrame: Data with startup/shutdown event flags
    """
    result_data = dispatch_data.copy()
    
    # Filter to only hydro generator units (hydro_t*), exclude pumps (hydro_p*)
    generator_units = [unit for unit in dispatch_data['unit_id'].unique() 
                      if 'hydro_t' in unit or ('hydro_' in unit and 'hydro_p' not in unit)]
    
    # Initialize event columns
    result_data['startup_event'] = 0
    result_data['shutdown_event'] = 0
    
    # Threshold for binary state comparison (handles floating point errors)
    threshold = 0.5
    
    # Group by unit_id to handle multiple hydro generator units
    for unit_id in generator_units:
        unit_mask = dispatch_data['unit_id'] == unit_id
        unit_data = dispatch_data[unit_mask].copy()
        
        # Sort by time to ensure proper order
        unit_data = unit_data.sort_values(['day', 'hour'])
        
        # Use u_G_jht for generator commitment
        commitment = unit_data['u_G_jht'].values
        
        # Initialize event arrays
        startup_events = np.zeros(len(unit_data))
        shutdown_events = np.zeros(len(unit_data))
        
        # Detect transitions using threshold comparison
        for i in range(1, len(commitment)):
            prev_state = 1 if commitment[i-1] > threshold else 0
            curr_state = 1 if commitment[i] > threshold else 0
            
            # Startup: 0 -> 1 transition
            if prev_state == 0 and curr_state == 1:
                startup_events[i] = 1
            # Shutdown: 1 -> 0 transition  
            elif prev_state == 1 and curr_state == 0:
                shutdown_events[i] = 1
        
        # Update the result dataframe
        result_data.loc[unit_mask, 'startup_event'] = startup_events
        result_data.loc[unit_mask, 'shutdown_event'] = shutdown_events
    
    return result_data


def detect_rough_zone_operation(dispatch_data: pd.DataFrame, rough_zone_segment_num: int, num_segments: int) -> pd.DataFrame:
    """
    Detect if each timestep is operating in the rough zone.
    
    Operating in rough zone IF:
    - Water flow exists in rough zone segment (q_G{N}_jht > threshold)
    - AND no water flow in next segment (q_G{N+1}_jht == 0 or doesn't exist)
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data
        rough_zone_segment_num (int): Segment number identified as rough zone
        num_segments (int): Total number of performance segments
        
    Returns:
        pd.DataFrame: Data with 'in_rough_zone' column added
    """
    result_data = dispatch_data.copy()
    
    # Initialize rough zone flag column
    result_data['in_rough_zone'] = False
    
    # Check if the required columns exist
    rough_zone_col = f'q_G{rough_zone_segment_num}_jht'
    
    if rough_zone_col not in result_data.columns:
        print(f"  Warning: Column '{rough_zone_col}' not found in dispatch data. Skipping rough zone detection.")
        return result_data
    
    # Determine if there's a next segment to check
    has_next_segment = rough_zone_segment_num < num_segments
    next_segment_col = f'q_G{rough_zone_segment_num + 1}_jht' if has_next_segment else None
    
    # Threshold for considering water flow as non-zero (handle floating point)
    threshold = 1e-6
    
    # Group by unit_id to handle multiple hydro units
    for unit_id in dispatch_data['unit_id'].unique():
        unit_mask = dispatch_data['unit_id'] == unit_id
        
        # Check if operating in rough zone segment
        in_rough_segment = result_data.loc[unit_mask, rough_zone_col] > threshold
        
        # Check if NOT operating beyond rough zone
        if has_next_segment and next_segment_col in result_data.columns:
            not_beyond_rough = result_data.loc[unit_mask, next_segment_col] <= threshold
        else:
            # If there's no next segment, we can't be beyond it
            not_beyond_rough = True
        
        # In rough zone = in rough segment AND not beyond it
        result_data.loc[unit_mask, 'in_rough_zone'] = in_rough_segment & not_beyond_rough
    
    return result_data


def calculate_hydro_generator_ramping(dispatch_data: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate ramping up and down amounts for each time step.
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data
        
    Returns:
        pd.DataFrame: Data with ramping amounts
    """
    result_data = dispatch_data.copy()
    
    # Initialize ramping columns
    result_data['ramp_up_mw'] = 0.0
    result_data['ramp_down_mw'] = 0.0
    
    # Group by unit_id to handle multiple hydro units
    for unit_id in dispatch_data['unit_id'].unique():
        unit_mask = dispatch_data['unit_id'] == unit_id
        unit_data = dispatch_data[unit_mask].copy()
        
        # Sort by time to ensure proper order
        unit_data = unit_data.sort_values(['day', 'hour'])
        
        # Calculate power differences
        power = unit_data['p_G_jht'].values
        
        # Initialize ramping arrays
        ramp_up_mw = np.zeros(len(unit_data))
        ramp_down_mw = np.zeros(len(unit_data))
        
        # Calculate ramping for each time step
        for i in range(1, len(power)):
            power_diff = power[i] - power[i-1]
            
            if power_diff > 0:
                ramp_up_mw[i] = power_diff
            elif power_diff < 0:
                ramp_down_mw[i] = abs(power_diff)
        
        # Update the result dataframe
        result_data.loc[unit_mask, 'ramp_up_mw'] = ramp_up_mw
        result_data.loc[unit_mask, 'ramp_down_mw'] = ramp_down_mw
    
    return result_data


def calculate_hourly_om_costs(dispatch_data: pd.DataFrame, cost_rates: Dict[str, Dict[str, float]], 
                               rough_zone_flag: bool = False) -> pd.DataFrame:
    """
    Calculate hourly O&M costs for each time step based on Tech_ID-specific cost rates.
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data with events and ramping
        cost_rates (Dict[str, Dict[str, float]]): O&M cost rates per Tech_ID
        rough_zone_flag (bool): Whether to include rough zone costs
        
    Returns:
        pd.DataFrame: Data with hourly O&M costs
    """
    result_data = dispatch_data.copy()
    
    # Initialize cost columns
    result_data['startup_cost_$'] = 0.0
    result_data['shutdown_cost_$'] = 0.0
    result_data['ramp_up_cost_$'] = 0.0
    result_data['ramp_down_cost_$'] = 0.0
    result_data['rough_zone_cost_$'] = 0.0
    result_data['total_hourly_om_cost_$'] = 0.0
    
    # Calculate costs for each row based on its Tech_ID
    for unit_id in dispatch_data['unit_id'].unique():
        if unit_id in cost_rates:
            unit_costs = cost_rates[unit_id]
            unit_mask = dispatch_data['unit_id'] == unit_id
            
            # Calculate costs using Tech_ID-specific rates
            startup_costs = result_data.loc[unit_mask, 'startup_event'] * unit_costs.get('startup_cost', 0)
            shutdown_costs = result_data.loc[unit_mask, 'shutdown_event'] * unit_costs.get('shutdown_cost', 0)
            ramp_up_costs = result_data.loc[unit_mask, 'ramp_up_mw'] * unit_costs.get('ramp_up_cost', 0)
            ramp_down_costs = result_data.loc[unit_mask, 'ramp_down_mw'] * unit_costs.get('ramp_down_cost', 0)
            
            # Calculate rough zone costs if enabled and column exists
            rough_zone_costs = 0.0
            if rough_zone_flag and 'in_rough_zone' in result_data.columns:
                rough_zone_costs = result_data.loc[unit_mask, 'in_rough_zone'].astype(float) * unit_costs.get('rough_zone_cost', 0)
            
            # Update cost columns
            result_data.loc[unit_mask, 'startup_cost_$'] = startup_costs
            result_data.loc[unit_mask, 'shutdown_cost_$'] = shutdown_costs
            result_data.loc[unit_mask, 'ramp_up_cost_$'] = ramp_up_costs
            result_data.loc[unit_mask, 'ramp_down_cost_$'] = ramp_down_costs
            result_data.loc[unit_mask, 'rough_zone_cost_$'] = rough_zone_costs
            result_data.loc[unit_mask, 'total_hourly_om_cost_$'] = (startup_costs + shutdown_costs + 
                                                                      ramp_up_costs + ramp_down_costs + 
                                                                      rough_zone_costs)
    
    return result_data


def generate_per_hydro_generator_cost_summary(dispatch_data: pd.DataFrame, rough_zone_flag: bool = False) -> Dict[str, Dict[str, float]]:
    """
    Generate per-unit O&M cost summary for each hydro generator.
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data with calculated costs
        rough_zone_flag (bool): Whether to include rough zone costs
        
    Returns:
        Dict[str, Dict[str, float]]: Per-unit summary of annual costs
    """
    per_unit_summary = {}
    
    # Filter to only hydro generator units (hydro_t*), exclude pumps (hydro_p*)
    generator_units = [unit for unit in dispatch_data['unit_id'].unique() 
                      if 'hydro_t' in unit or ('hydro_' in unit and 'hydro_p' not in unit)]
    
    for unit_id in generator_units:
        unit_data = dispatch_data[dispatch_data['unit_id'] == unit_id]
        
        per_unit_summary[unit_id] = {
            'startup_events': unit_data['startup_event'].sum(),
            'shutdown_events': unit_data['shutdown_event'].sum(),
            'ramp_up_mw': unit_data['ramp_up_mw'].sum(),
            'ramp_down_mw': unit_data['ramp_down_mw'].sum(),
            'startup_cost_$': unit_data['startup_cost_$'].sum(),
            'shutdown_cost_$': unit_data['shutdown_cost_$'].sum(),
            'ramp_up_cost_$': unit_data['ramp_up_cost_$'].sum(),
            'ramp_down_cost_$': unit_data['ramp_down_cost_$'].sum(),
            'total_om_cost_$': unit_data['total_hourly_om_cost_$'].sum()
        }
        
        # Add rough zone data if enabled
        if rough_zone_flag and 'in_rough_zone' in unit_data.columns:
            per_unit_summary[unit_id]['rough_zone_hours'] = unit_data['in_rough_zone'].sum()
            per_unit_summary[unit_id]['rough_zone_cost_$'] = unit_data['rough_zone_cost_$'].sum()
    
    return per_unit_summary


def calculate_hydro_pump_startup_shutdown_events(dispatch_data: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate startup and shutdown events for pump units based on commitment transitions.
    Only processes pump units (hydro_p*).
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data
        
    Returns:
        pd.DataFrame: Data with pump startup/shutdown event flags
    """
    result_data = dispatch_data.copy()
    
    # Filter to only pump units (hydro_p*)
    pump_units = [unit for unit in dispatch_data['unit_id'].unique() if 'hydro_p' in unit]
    
    # Initialize event columns if not already present
    if 'startup_event' not in result_data.columns:
        result_data['startup_event'] = 0
    if 'shutdown_event' not in result_data.columns:
        result_data['shutdown_event'] = 0
    
    # Threshold for binary state comparison (handles floating point errors)
    threshold = 0.5
    
    # Group by unit_id to handle multiple pump units
    for unit_id in pump_units:
        unit_mask = dispatch_data['unit_id'] == unit_id
        unit_data = dispatch_data[unit_mask].copy()
        
        # Sort by time to ensure proper order
        unit_data = unit_data.sort_values(['day', 'hour'])
        
        # Use u_P_jht for pump commitment
        if 'u_P_jht' in unit_data.columns:
            commitment = unit_data['u_P_jht'].values
            
            # Initialize event arrays
            startup_events = np.zeros(len(unit_data))
            shutdown_events = np.zeros(len(unit_data))
            
            # Detect transitions using threshold comparison
            for i in range(1, len(commitment)):
                prev_state = 1 if commitment[i-1] > threshold else 0
                curr_state = 1 if commitment[i] > threshold else 0
                
                # Startup: 0 -> 1 transition
                if prev_state == 0 and curr_state == 1:
                    startup_events[i] = 1
                # Shutdown: 1 -> 0 transition  
                elif prev_state == 1 and curr_state == 0:
                    shutdown_events[i] = 1
            
            # Update the result dataframe
            result_data.loc[unit_mask, 'startup_event'] = startup_events
            result_data.loc[unit_mask, 'shutdown_event'] = shutdown_events
    
    return result_data


def calculate_hydro_pump_ramping(dispatch_data: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate ramping up and down amounts for pump units.
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data
        
    Returns:
        pd.DataFrame: Data with pump ramping amounts
    """
    result_data = dispatch_data.copy()
    
    # Initialize ramping columns if not already present
    if 'ramp_up_mw' not in result_data.columns:
        result_data['ramp_up_mw'] = 0.0
    if 'ramp_down_mw' not in result_data.columns:
        result_data['ramp_down_mw'] = 0.0
    
    # Filter to only pump units (hydro_p*)
    pump_units = [unit for unit in dispatch_data['unit_id'].unique() if 'hydro_p' in unit]
    
    # Group by unit_id to handle multiple pump units
    for unit_id in pump_units:
        unit_mask = dispatch_data['unit_id'] == unit_id
        unit_data = dispatch_data[unit_mask].copy()
        
        # Sort by time to ensure proper order
        unit_data = unit_data.sort_values(['day', 'hour'])
        
        # Use p_P_jht for pump power
        if 'p_P_jht' in unit_data.columns:
            power = unit_data['p_P_jht'].values
            
            # Initialize ramping arrays
            ramp_up_mw = np.zeros(len(unit_data))
            ramp_down_mw = np.zeros(len(unit_data))
            
            # Calculate ramping for each time step
            for i in range(1, len(power)):
                power_diff = power[i] - power[i-1]
                
                if power_diff > 0:
                    ramp_up_mw[i] = power_diff
                elif power_diff < 0:
                    ramp_down_mw[i] = abs(power_diff)
            
            # Update the result dataframe
            result_data.loc[unit_mask, 'ramp_up_mw'] = ramp_up_mw
            result_data.loc[unit_mask, 'ramp_down_mw'] = ramp_down_mw
    
    return result_data


def generate_per_hydro_pump_cost_summary(dispatch_data: pd.DataFrame, rough_zone_flag: bool = False) -> Dict[str, Dict[str, float]]:
    """
    Generate per-pump O&M cost summary for each pump unit.
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data with calculated costs
        rough_zone_flag (bool): Whether to include rough zone costs
        
    Returns:
        Dict[str, Dict[str, float]]: Per-pump summary of annual costs
    """
    per_pump_summary = {}
    
    # Filter to only pump units (hydro_p*)
    pump_units = [unit for unit in dispatch_data['unit_id'].unique() if 'hydro_p' in unit]
    
    for unit_id in pump_units:
        unit_data = dispatch_data[dispatch_data['unit_id'] == unit_id]
        
        per_pump_summary[unit_id] = {
            'startup_events': unit_data['startup_event'].sum(),
            'shutdown_events': unit_data['shutdown_event'].sum(),
            'ramp_up_mw': unit_data['ramp_up_mw'].sum(),
            'ramp_down_mw': unit_data['ramp_down_mw'].sum(),
            'startup_cost_$': unit_data['startup_cost_$'].sum(),
            'shutdown_cost_$': unit_data['shutdown_cost_$'].sum(),
            'ramp_up_cost_$': unit_data['ramp_up_cost_$'].sum(),
            'ramp_down_cost_$': unit_data['ramp_down_cost_$'].sum(),
            'total_om_cost_$': unit_data['total_hourly_om_cost_$'].sum()
        }
        
        # Add rough zone data if enabled
        if rough_zone_flag and 'in_rough_zone' in unit_data.columns:
            per_pump_summary[unit_id]['rough_zone_hours'] = unit_data['in_rough_zone'].sum()
            per_pump_summary[unit_id]['rough_zone_cost_$'] = unit_data['rough_zone_cost_$'].sum()
    
    return per_pump_summary


def generate_om_cost_summary(dispatch_data: pd.DataFrame, rough_zone_flag: bool = False) -> Dict[str, float]:
    """
    Generate annual O&M cost summary.
    
    Args:
        dispatch_data (pd.DataFrame): Hydro dispatch data with calculated costs
        rough_zone_flag (bool): Whether to include rough zone costs
        
    Returns:
        Dict[str, float]: Summary of annual costs
    """
    summary = {
        'total_startup_events': dispatch_data['startup_event'].sum(),
        'total_shutdown_events': dispatch_data['shutdown_event'].sum(),
        'total_ramp_up_mw': dispatch_data['ramp_up_mw'].sum(),
        'total_ramp_down_mw': dispatch_data['ramp_down_mw'].sum(),
        'annual_startup_cost_$': dispatch_data['startup_cost_$'].sum(),
        'annual_shutdown_cost_$': dispatch_data['shutdown_cost_$'].sum(),
        'annual_ramp_up_cost_$': dispatch_data['ramp_up_cost_$'].sum(),
        'annual_ramp_down_cost_$': dispatch_data['ramp_down_cost_$'].sum(),
        'total_annual_om_cost_$': dispatch_data['total_hourly_om_cost_$'].sum()
    }
    
    # Add rough zone data if enabled
    if rough_zone_flag and 'in_rough_zone' in dispatch_data.columns:
        summary['total_rough_zone_hours'] = dispatch_data['in_rough_zone'].sum()
        summary['annual_rough_zone_cost_$'] = dispatch_data['rough_zone_cost_$'].sum()
    
    return summary


def display_cost_rates(cost_rates: Dict[str, Dict[str, float]]):
    """
    Display O&M cost rates for each Tech_ID.
    
    Args:
        cost_rates (Dict[str, Dict[str, float]]): Cost rates per Tech_ID
    """
    print("\nO&M COST RATES:")
    print("-"*70)
    
    for tech_id in sorted(cost_rates.keys()):
        print(f"\n{tech_id}:")
        for cost_type, value in sorted(cost_rates[tech_id].items()):
            print(f"  {cost_type}: ${value:,.4f}")


def display_om_cost_summary(summary: Dict[str, float], model_type: str):
    """
    Display the annual O&M cost summary.
    
    Args:
        summary (Dict[str, float]): Cost summary data
        model_type (str): Model type description
    """
    print("\n" + "="*70)
    print(f"ANNUAL HYDRO O&M COST SUMMARY - {model_type}")
    print("="*70)
    
    print(f"\nOPERATIONAL EVENTS:")
    print(f"  Total Startup Events: {summary['total_startup_events']:.0f}")
    print(f"  Total Shutdown Events: {summary['total_shutdown_events']:.0f}")
    print(f"  Total Ramp Up (MW): {summary['total_ramp_up_mw']:.2f}")
    print(f"  Total Ramp Down (MW): {summary['total_ramp_down_mw']:.2f}")
    if 'total_rough_zone_hours' in summary:
        print(f"  Total Rough Zone Hours: {summary['total_rough_zone_hours']:.0f}")
    
    print(f"\nANNUAL COSTS:")
    print(f"  Startup Costs: ${summary['annual_startup_cost_$']:,.2f}")
    print(f"  Shutdown Costs: ${summary['annual_shutdown_cost_$']:,.2f}")
    print(f"  Ramp Up Costs: ${summary['annual_ramp_up_cost_$']:,.2f}")
    print(f"  Ramp Down Costs: ${summary['annual_ramp_down_cost_$']:,.2f}")
    if 'annual_rough_zone_cost_$' in summary:
        print(f"  Rough Zone Costs: ${summary['annual_rough_zone_cost_$']:,.2f}")
    print(f"  TOTAL ANNUAL O&M COST: ${summary['total_annual_om_cost_$']:,.2f}")
    
    print("="*70)


def display_per_hydro_generator_cost_summary(per_unit_summary: Dict[str, Dict[str, float]], model_type: str):
    """
    Display the per-unit O&M cost breakdown.
    
    Args:
        per_unit_summary (Dict[str, Dict[str, float]]): Per-unit cost summary data
        model_type (str): Model type description
    """
    print("\n" + "="*80)
    print(f"PER-UNIT HYDRO O&M COST BREAKDOWN - {model_type}")
    print("="*80)
    
    for unit_id in sorted(per_unit_summary.keys()):
        unit_costs = per_unit_summary[unit_id]
        print(f"\n{unit_id.upper()}:")
        print("-" * 40)
        print(f"  Startup Events: {unit_costs['startup_events']:.0f}")
        print(f"  Shutdown Events: {unit_costs['shutdown_events']:.0f}")
        print(f"  Ramp Up (MW): {unit_costs['ramp_up_mw']:.2f}")
        print(f"  Ramp Down (MW): {unit_costs['ramp_down_mw']:.2f}")
        if 'rough_zone_hours' in unit_costs:
            print(f"  Rough Zone Hours: {unit_costs['rough_zone_hours']:.0f}")
        print(f"  Startup Costs: ${unit_costs['startup_cost_$']:,.2f}")
        print(f"  Shutdown Costs: ${unit_costs['shutdown_cost_$']:,.2f}")
        print(f"  Ramp Up Costs: ${unit_costs['ramp_up_cost_$']:,.2f}")
        print(f"  Ramp Down Costs: ${unit_costs['ramp_down_cost_$']:,.2f}")
        if 'rough_zone_cost_$' in unit_costs:
            print(f"  Rough Zone Costs: ${unit_costs['rough_zone_cost_$']:,.2f}")
        print(f"  UNIT TOTAL O&M COST: ${unit_costs['total_om_cost_$']:,.2f}")
    
    print("="*80)


def display_per_hydro_pump_cost_summary(per_pump_summary: Dict[str, Dict[str, float]], model_type: str):
    """
    Display the per-pump O&M cost breakdown.
    
    Args:
        per_pump_summary (Dict[str, Dict[str, float]]): Per-pump cost summary data
        model_type (str): Model type description
    """
    print("\n" + "="*80)
    print(f"PER-PUMP O&M COST BREAKDOWN - {model_type}")
    print("="*80)
    
    for unit_id in sorted(per_pump_summary.keys()):
        unit_costs = per_pump_summary[unit_id]
        print(f"\n{unit_id.upper()}:")
        print("-" * 40)
        print(f"  Startup Events: {unit_costs['startup_events']:.0f}")
        print(f"  Shutdown Events: {unit_costs['shutdown_events']:.0f}")
        print(f"  Ramp Up (MW): {unit_costs['ramp_up_mw']:.2f}")
        print(f"  Ramp Down (MW): {unit_costs['ramp_down_mw']:.2f}")
        if 'rough_zone_hours' in unit_costs:
            print(f"  Rough Zone Hours: {unit_costs['rough_zone_hours']:.0f}")
        print(f"  Startup Costs: ${unit_costs['startup_cost_$']:,.2f}")
        print(f"  Shutdown Costs: ${unit_costs['shutdown_cost_$']:,.2f}")
        print(f"  Ramp Up Costs: ${unit_costs['ramp_up_cost_$']:,.2f}")
        print(f"  Ramp Down Costs: ${unit_costs['ramp_down_cost_$']:,.2f}")
        if 'rough_zone_cost_$' in unit_costs:
            print(f"  Rough Zone Costs: ${unit_costs['rough_zone_cost_$']:,.2f}")
        print(f"  PUMP TOTAL O&M COST: ${unit_costs['total_om_cost_$']:,.2f}")
    
    print("="*80)


def create_om_cost_csv(dispatch_data: pd.DataFrame, output_filepath: str, rough_zone_flag: bool = False):
    """
    Create a separate O&M cost CSV file with time series data.
    One row per timestep (day/hour), with detailed cost columns for each unit.
    
    For each unit, includes:
    - startup_event_{unit_id}
    - shutdown_event_{unit_id}
    - ramp_up_mw_{unit_id}
    - ramp_down_mw_{unit_id}
    - startup_cost_$_{unit_id}
    - shutdown_cost_$_{unit_id}
    - ramp_up_cost_$_{unit_id}
    - ramp_down_cost_$_{unit_id}
    - rough_zone_cost_$_{unit_id} (if rough_zone_flag is True)
    - O&M_{unit_id}
    
    Plus O&M_Total (sum of all unit O&M costs).
    
    Args:
        dispatch_data (pd.DataFrame): Dispatch data with calculated O&M costs
        output_filepath (str): Path where the O&M cost CSV will be saved
        rough_zone_flag (bool): Whether to include rough zone cost columns
    """
    # Get unique timesteps
    timesteps = dispatch_data[['day', 'hour']].drop_duplicates().sort_values(['day', 'hour']).reset_index(drop=True)
    
    # Get unique unit_ids (sorted)
    unit_ids = sorted(dispatch_data['unit_id'].unique())
    
    # For each unit, create detail columns
    for unit_id in unit_ids:
        unit_data = dispatch_data[dispatch_data['unit_id'] == unit_id].copy()
        
        # Define the columns we want to extract
        detail_columns = {
            f'startup_event_{unit_id}': 'startup_event',
            f'shutdown_event_{unit_id}': 'shutdown_event',
            f'ramp_up_mw_{unit_id}': 'ramp_up_mw',
            f'ramp_down_mw_{unit_id}': 'ramp_down_mw',
            f'startup_cost_$_{unit_id}': 'startup_cost_$',
            f'shutdown_cost_$_{unit_id}': 'shutdown_cost_$',
            f'ramp_up_cost_$_{unit_id}': 'ramp_up_cost_$',
            f'ramp_down_cost_$_{unit_id}': 'ramp_down_cost_$'
        }
        
        # Add rough zone cost column if enabled
        if rough_zone_flag and 'rough_zone_cost_$' in unit_data.columns:
            detail_columns[f'rough_zone_cost_$_{unit_id}'] = 'rough_zone_cost_$'
        
        # Always add O&M total for the unit
        detail_columns[f'O&M_{unit_id}'] = 'total_hourly_om_cost_$'
        
        # Aggregate by day and hour
        for new_col_name, orig_col_name in detail_columns.items():
            if orig_col_name in unit_data.columns:
                unit_detail = unit_data.groupby(['day', 'hour'])[orig_col_name].sum().reset_index()
                unit_detail = unit_detail.rename(columns={orig_col_name: new_col_name})
                
                # Merge with timesteps
                timesteps = timesteps.merge(unit_detail, on=['day', 'hour'], how='left')
                timesteps[new_col_name] = timesteps[new_col_name].fillna(0.0)
    
    # Create O&M_Total column (sum across all units for each timestep)
    om_columns = [col for col in timesteps.columns if col.startswith('O&M_') and col != 'O&M_Total']
    timesteps['O&M_Total'] = timesteps[om_columns].sum(axis=1)
    
    # Round all numeric columns to eliminate floating point precision errors
    # Set values below threshold (1e-10) to exactly 0
    numeric_columns = timesteps.select_dtypes(include=[np.number]).columns
    for col in numeric_columns:
        timesteps[col] = timesteps[col].round(10)
        timesteps.loc[abs(timesteps[col]) < 1e-10, col] = 0.0
    
    # Save to CSV
    timesteps.to_csv(output_filepath, index=False)


def print_validation_results(model_type: str, hydro_valid: bool, hydro_missing: List[str],
                            hydro_count: int, reservoir_info: List[Tuple[str, float, float]]):
    """
    Print formatted validation results.
    
    Args:
        model_type (str): Detected model type
        hydro_valid (bool): Whether hydro validation passed
        hydro_missing (List[str]): List of missing hydro Tech_IDs
        hydro_count (int): Total number of hydro Tech_IDs
        reservoir_info (List[Tuple[str, float, float]]): Reservoir information
    """
    print("\n" + "="*70)
    print(f"MODEL TYPE: {model_type}")
    print("="*70)
    
    # Print reservoir information if available
    if reservoir_info:
        print("\nRESERVOIR INFORMATION:")
        print("-"*70)
        for res_id, min_vol, max_vol in reservoir_info:
            print(f"  {res_id}:")
            print(f"    Min Volume: {min_vol:,.2f} acre-feet")
            print(f"    Max Volume: {max_vol:,.2f} acre-feet")
        print("-"*70)
    
    print("\nVALIDATION RESULTS:")
    print("-"*70)
    
    # Hydro validation
    if hydro_count > 0:
        if hydro_valid:
            print(f"✓ Hydro: All {hydro_count} Tech_IDs defined in O&M Cost sheet")
        else:
            print(f"✗ Hydro: Missing O&M cost definitions for {len(hydro_missing)} Tech_ID(s):")
            for tech_id in hydro_missing:
                print(f"    - {tech_id}")
    else:
        print("○ Hydro: No Tech_IDs found in Gen Technology - Hydro sheet")
    
    # Overall result
    print("-"*70)
    if hydro_valid:
        print("✓ ALL VALIDATIONS PASSED!")
    else:
        print("✗ VALIDATION FAILED - Please add missing Tech_IDs to O&M Cost sheet")
    
    print("="*70 + "\n")


def main():
    """
    Main function to orchestrate model type detection, validation, and O&M cost calculation.
    """
    # Parse arguments
    filename = parse_CLI_arguments()
    
    # Construct full filepath
    filepath = construct_filepath(filename)
    
    # Check if file exists
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)
    
    print(f"\nAnalyzing: {filepath}")
    
    # Detect model type
    model_type = detect_model_type(filepath)
    
    # Get reservoir information
    reservoir_info = get_reservoir_info(filepath)
    
    # Extract Tech_IDs from Gen Technology sheets
    hydro_ids = get_gen_tech_hydro_ids(filepath)
    
    # Extract Tech_IDs from O&M Cost sheet
    om_cost_ids = get_om_cost_tech_ids(filepath)
    
    # Validate
    hydro_valid, hydro_missing = validate_tech_ids(hydro_ids, om_cost_ids, "Hydro")
    
    # Print validation results
    print_validation_results(
        model_type, 
        hydro_valid, hydro_missing, len(hydro_ids),
        reservoir_info
    )
    
    # Exit if validation failed
    if not hydro_valid:
        sys.exit(1)
    
    # Extract O&M cost rates
    print("\nExtracting O&M cost rates...")
    cost_rates = get_om_cost_rates_per_tech_id(filepath)
    
    if cost_rates:
        display_cost_rates(cost_rates)
    else:
        print("Warning: No cost rates extracted")
        sys.exit(1)
    
    # Construct dispatch file path
    dispatch_filepath = construct_dispatch_filepath(filename)
    
    # Check if dispatch file exists
    if not os.path.exists(dispatch_filepath):
        print(f"\n{'='*70}")
        print("DISPATCH FILE NOT FOUND")
        print(f"{'='*70}")
        print(f"Expected path: {dispatch_filepath}")
        print("\nO&M cost rates have been validated and extracted.")
        print("Run simulation to generate dispatch data for cost calculations.")
        print(f"{'='*70}\n")
        return
    
    # Get rough zone settings
    print("\nChecking rough zone settings...")
    rough_zone_flag, rough_zone_segment_num, num_segments = get_rough_zone_settings(filepath)
    
    if rough_zone_flag:
        print(f"  Rough zone operation enabled")
        print(f"  Rough zone segment: {rough_zone_segment_num}")
        print(f"  Total segments: {num_segments}")
    else:
        print(f"  Rough zone operation disabled")
    
    # Import and process dispatch data
    print(f"\nProcessing dispatch data from: {dispatch_filepath}")
    try:
        # Import dispatch data
        dispatch_data = import_hydro_dispatch_data(dispatch_filepath)
        
        # Calculate startup/shutdown events
        print("  Calculating startup/shutdown events...")
        dispatch_data = calculate_hydro_generator_startup_shutdown_events(dispatch_data)
        
        # Calculate ramping
        print("  Calculating ramping...")
        dispatch_data = calculate_hydro_generator_ramping(dispatch_data)
        
        # Detect rough zone operation if enabled
        if rough_zone_flag:
            print(f"  Detecting rough zone operation...")
            dispatch_data = detect_rough_zone_operation(dispatch_data, rough_zone_segment_num, num_segments)
        else:
            print(f"  Rough zone operation not active - skipping rough zone cost calculation")
        
        # Calculate hourly O&M costs
        print("  Calculating hourly O&M costs...")
        dispatch_data = calculate_hourly_om_costs(dispatch_data, cost_rates, rough_zone_flag)
        
        # Check if this is a pumped storage model and process pumps
        is_pumped_storage = (model_type == "Pumped Storage")
        
        if is_pumped_storage:
            # Check if we have pump units
            pump_units = [unit for unit in dispatch_data['unit_id'].unique() if 'hydro_p' in unit]
            
            if pump_units:
                print("  Calculating pump startup/shutdown events...")
                dispatch_data = calculate_hydro_pump_startup_shutdown_events(dispatch_data)
                
                print("  Calculating pump ramping...")
                dispatch_data = calculate_hydro_pump_ramping(dispatch_data)
                
                print("  Calculating pump hourly O&M costs...")
                dispatch_data = calculate_hourly_om_costs(dispatch_data, cost_rates, rough_zone_flag)
        
        # Generate summaries
        print("  Generating cost summaries...")
        summary = generate_om_cost_summary(dispatch_data, rough_zone_flag)
        per_unit_summary = generate_per_hydro_generator_cost_summary(dispatch_data, rough_zone_flag)
        
        # Display results
        display_om_cost_summary(summary, model_type)
        display_per_hydro_generator_cost_summary(per_unit_summary, model_type)
        
        # Display pump results if pumped storage
        if is_pumped_storage:
            pump_units = [unit for unit in dispatch_data['unit_id'].unique() if 'hydro_p' in unit]
            if pump_units:
                per_pump_summary = generate_per_hydro_pump_cost_summary(dispatch_data, rough_zone_flag)
                display_per_hydro_pump_cost_summary(per_pump_summary, model_type)
        
        # Create separate O&M cost CSV file
        om_cost_filepath = os.path.join("Simulation_Results", filename, f"ALEAF_HydroBoost_{filename}__OM_cost.csv")
        print(f"\n  Creating O&M cost CSV file: {om_cost_filepath}")
        create_om_cost_csv(dispatch_data, om_cost_filepath, rough_zone_flag)
        
        print("\n✓ O&M cost calculation completed successfully!")
        print(f"✓ O&M cost time series saved to: {om_cost_filepath}\n")
        
    except Exception as e:
        print(f"\nError during O&M cost calculation: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
